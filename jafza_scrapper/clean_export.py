"""
Produces a clean, recruiter-friendly XLSX from the enriched output --
just company identity + best available address, no confidence scores,
match reasoning, or any of the internal diagnostic columns.

Usage:
    python clean_export.py --input output/jafza_google_enriched.xlsx --output output/jafza_companies_clean.xlsx
"""
from __future__ import annotations

import argparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="output/jafza_google_enriched.xlsx")
    parser.add_argument("--output", default="output/jafza_companies_clean.xlsx")
    args = parser.parse_args()

    src_wb = load_workbook(args.input, data_only=True)
    src_ws = src_wb.active
    headers = [c.value for c in next(src_ws.iter_rows(min_row=1, max_row=1))]

    lower = [str(h).strip().lower() if h else "" for h in headers]

    def find(*names):
        for n in names:
            if n in lower:
                return lower.index(n)
        return None

    sr_col = find("sr no")
    company_col = find("company")
    email_col = find("email")
    phone_col = find("contact no")
    google_addr_col = find("google_formatted_address")
    original_addr_col = find("address")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Companies"
    out_ws.append(["SR NO", "Company", "Email", "Contact No", "Address"])

    written = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        sr = row[sr_col] if sr_col is not None else ""
        company = row[company_col] if company_col is not None else ""
        email = row[email_col] if email_col is not None else ""
        phone = row[phone_col] if phone_col is not None else ""

        # Prefer the Google-resolved address; fall back to whatever was
        # already in the original Address column, if anything.
        address = ""
        if google_addr_col is not None and row[google_addr_col]:
            address = row[google_addr_col]
        elif original_addr_col is not None and row[original_addr_col]:
            address = row[original_addr_col]

        out_ws.append([sr, company, email or "", phone or "", address])
        written += 1

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in out_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    out_ws.row_dimensions[1].height = 24

    for row in out_ws.iter_rows(min_row=2, max_row=out_ws.max_row):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border

    widths = [8, 42, 32, 16, 55]
    for i, w in enumerate(widths, start=1):
        out_ws.column_dimensions[out_ws.cell(row=1, column=i).column_letter].width = w
    out_ws.freeze_panes = "A2"

    out_wb.save(args.output)
    print(f"Clean export written: {args.output} ({written} companies)")


if __name__ == "__main__":
    main()