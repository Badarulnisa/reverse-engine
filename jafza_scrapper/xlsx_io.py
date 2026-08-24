"""
XLSX input/output for the enrichment pipeline.

Guarantees:
- The source workbook is opened read-only conceptually: we load it,
  copy every original column verbatim into the output, and never
  mutate a source cell.
- Enrichment columns are appended after the original columns, never
  interleaved or overwritten.
- Row order / row identity is preserved 1:1 (source row N -> output
  row N), so every enriched row stays traceable back to its source row.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from matcher import MatchResult

ENRICHMENT_HEADERS = [
    "google_match_status",
    "google_match_confidence",
    "google_place_id",
    "google_name",
    "google_formatted_address",
    "google_latitude",
    "google_longitude",
    "google_maps_url",
    "google_website",
    "google_phone",
    "google_business_types",
    "google_match_reasoning",
]


@dataclass
class SourceRow:
    row_index: int              # 1-based row number in the ORIGINAL sheet (traceability)
    values: list                # raw cell values, original column order, untouched
    company_name: str
    email: Optional[str]
    phone: Optional[str]


def detect_company_name_column(headers: list[str]) -> Optional[int]:
    """Returns 0-based column index, or None if not confidently detected."""
    candidates = ("company", "company name", "name", "business name")
    lowered = [str(h).strip().lower() if h else "" for h in headers]
    for cand in candidates:
        if cand in lowered:
            return lowered.index(cand)
    return None


def load_source(path: str, company_col: Optional[int] = None,
                 email_col: Optional[int] = None,
                 phone_col: Optional[int] = None) -> tuple[list[str], list[SourceRow]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    if company_col is None:
        company_col = detect_company_name_column(headers)
        if company_col is None:
            raise ValueError(
                f"Could not auto-detect the company-name column from headers "
                f"{headers!r}. Pass company_col explicitly."
            )

    if email_col is None:
        lowered = [str(h).strip().lower() if h else "" for h in headers]
        email_col = lowered.index("email") if "email" in lowered else None

    if phone_col is None:
        lowered = [str(h).strip().lower() if h else "" for h in headers]
        for cand in ("contact no", "phone", "contact number"):
            if cand in lowered:
                phone_col = lowered.index(cand)
                break

    rows: list[SourceRow] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[company_col]
        if name is None or str(name).strip() == "":
            continue
        email = row[email_col] if email_col is not None and email_col < len(row) else None
        phone = row[phone_col] if phone_col is not None and phone_col < len(row) else None
        rows.append(SourceRow(
            row_index=i,
            values=list(row),
            company_name=str(name).strip(),
            email=(str(email).strip() if email else None),
            phone=(str(phone).strip() if phone else None),
        ))

    return headers, rows


def _style_header(ws) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _style_body(ws, n_cols: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border


def write_enriched(
    output_path: str,
    original_headers: list[str],
    rows: list[SourceRow],
    results: dict[int, MatchResult],   # keyed by SourceRow.row_index
) -> None:
    """Writes the full enriched workbook: every original row + appended
    enrichment columns. Unresolved rows are included too (never
    silently dropped), with status/confidence making that explicit."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched"

    ws.append(list(original_headers) + ENRICHMENT_HEADERS)

    for r in rows:
        result = results.get(r.row_index)
        enrichment = _result_to_row(result)
        ws.append(list(r.values) + enrichment)

    _style_header(ws)
    _style_body(ws, len(original_headers) + len(ENRICHMENT_HEADERS))

    for i in range(1, len(original_headers) + len(ENRICHMENT_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 24
    # company/address-ish columns wider
    ws.column_dimensions["B"].width = 40

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_unresolved(
    output_path: str,
    original_headers: list[str],
    rows: list[SourceRow],
    results: dict[int, MatchResult],
) -> int:
    """Separate review file containing only rows that need a human
    look: unresolved, low-confidence, or API errors. Returns count."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Needs Review"

    review_headers = list(original_headers) + [
        "google_match_status", "google_match_confidence",
        "google_match_reasoning", "google_candidates_considered", "google_error_message",
        "top_candidate_name", "top_candidate_address",
        "top_candidate_score", "top_candidate_reasoning",
    ]
    ws.append(review_headers)

    count = 0
    for r in rows:
        result = results.get(r.row_index)
        if result is None:
            continue
        needs_review = (
            result.status in ("unresolved", "api_error")
            or result.confidence in ("low", "none")
        )
        if not needs_review:
            continue
        ws.append(list(r.values) + [
            result.status,
            result.confidence,
            result.reasoning,
            result.candidates_considered,
            result.error_message or "",
            result.top_candidate_name or "",
            result.top_candidate_address or "",
            result.top_candidate_score if result.top_candidate_score is not None else "",
            result.top_candidate_reasoning or "",
        ])
        count += 1

    _style_header(ws)
    _style_body(ws, len(review_headers))
    for i in range(1, len(review_headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 24
    ws.column_dimensions["B"].width = 40

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return count


def _result_to_row(result: Optional[MatchResult]) -> list:
    if result is None:
        return ["not_processed", "none", "", "", "", "", "", "", "", "", "", ""]
    return [
        result.status,
        result.confidence,
        result.place_id or "",
        result.name or "",
        result.formatted_address or "",
        result.latitude if result.latitude is not None else "",
        result.longitude if result.longitude is not None else "",
        result.maps_url or "",
        result.website or "",
        result.phone or "",
        result.business_types or "",
        result.reasoning or "",
    ]