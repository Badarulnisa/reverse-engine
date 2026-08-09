import os
import time
import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from app.core.har_parser import parse_har_file

print("[*] Parsing HAR file for session headers and endpoint template...")
har_path = "dmcc.ae.har"
records = parse_har_file(har_path)

target_rec = None
for rec in records:
    if "apexremote" in rec.get("url", "") and rec.get("method") == "POST":
        target_rec = rec
        break

if not target_rec:
    print("[!] Error: Could not find the Salesforce apexremote POST call in the HAR file.")
    exit(1)

url = target_rec.get("url")
raw_headers = target_rec.get("request_headers", {})

# Strip HTTP/2 pseudo-headers (":authority", ":method", ":path", ":scheme", etc.)
# and other headers that `requests` sets automatically or rejects outright.
DROP_HEADERS = {"content-length", "host"}
headers = {
    k: v for k, v in raw_headers.items()
    if not k.startswith(":") and k.lower() not in DROP_HEADERS
}

print(f"[+] Target Endpoint: {url}")
print(f"[+] Loaded {len(headers)} usable headers (dropped {len(raw_headers) - len(headers)} pseudo/reserved headers).")
print("[*] Configuring Tor proxy at socks5h://127.0.0.1:26000...")

session = requests.Session()
session.headers.update(headers)
session.proxies = {
    "http": "socks5h://127.0.0.1:26000",
    "https": "socks5h://127.0.0.1:26000"
}

PAGE_SIZE = 100
MAX_PAGES_PER_TERM = 300
CAP_PER_TERM = 1000

search_terms = list("abcdefghijklmnopqrstuvwxyz")

all_rows_by_license = {}

def fetch_term(term: str) -> int:
    added = 0
    for page in range(1, MAX_PAGES_PER_TERM + 1):
        payload = [
            {
                "action": "DMCCPublicDirectoryPage_Ctrl",
                "method": "searchCustomerDirectory",
                "data": [term, page, PAGE_SIZE],
                "type": "rpc",
                "tid": page
            }
        ]

        try:
            response = session.post(url, json=payload, timeout=60)
            if response.status_code != 200:
                print(f"    [!] HTTP Error {response.status_code} on term={term!r} page={page}")
                print(f"        Response body: {response.text[:300]}")
                break

            res_json = response.json()
            result_data = []
            if isinstance(res_json, list) and len(res_json) > 0:
                result_data = res_json[0].get("result", [])
            elif isinstance(res_json, dict):
                result_data = res_json.get("result", [])

            if not result_data:
                break

            for row in result_data:
                lic_no = row.get("licNo")
                key = lic_no or f"noLic_{row.get('customerName')}_{row.get('customerRegistrationNumber')}"
                if key in all_rows_by_license:
                    continue

                clean_row = {
                    "Company Name (English)": row.get("customerName"),
                    "Company Name (Arabic)": row.get("customerArabicName"),
                    "License Number": row.get("licNo"),
                    "License Issue Date": row.get("licIssueDate"),
                    "License Expiry Date": row.get("licExpDate"),
                    "License Address (English)": str(row.get("licAddress", "")).replace("<br>", "\n"),
                    "License Address (Arabic)": str(row.get("licArabicAddress", "")).replace("<br>", "\n"),
                    "License Manager": row.get("licManagerName"),
                    "License Manager (Arabic)": row.get("licManagerArabicName"),
                    "License Activity": str(row.get("licenseActivities", "")).replace("<br>", "\n"),
                    "License Activity (Arabic)": str(row.get("licenseArabicActivities", "")).replace("<br>", "\n"),
                    "License Status": row.get("licenseStatus"),
                    "Registration Status": row.get("customerRegistrationStatus"),
                    "Registration Number": row.get("customerRegistrationNumber"),
                    "Registration Date": row.get("customerRegistrationDate")
                }
                all_rows_by_license[key] = clean_row
                added += 1

            if len(result_data) < PAGE_SIZE:
                break

        except Exception as e:
            print(f"    [!] Exception on term={term!r} page={page}: {e}")
            break

        time.sleep(0.3)

    return added

print("[*] Starting per-letter extraction through Tor...\n")

terms_needing_split = []

for term in search_terms:
    added = fetch_term(term)
    total_now = len(all_rows_by_license)
    print(f"[+] term={term!r}: +{added} new rows (running total unique: {total_now})")

    if added >= CAP_PER_TERM:
        terms_needing_split.append(term)

    time.sleep(0.5)

if terms_needing_split:
    print(f"\n[*] These terms appear capped and need finer splitting: {terms_needing_split}")
    print("[*] Expanding into two-letter prefixes for those...\n")

    import string
    for base in terms_needing_split:
        for second in string.ascii_lowercase:
            term2 = base + second
            added = fetch_term(term2)
            total_now = len(all_rows_by_license)
            if added > 0:
                print(f"[+] term={term2!r}: +{added} new rows (running total unique: {total_now})")
            time.sleep(0.3)

print(f"\n[*] Extraction complete. Total unique records: {len(all_rows_by_license)}")

all_extracted_rows = list(all_rows_by_license.values())

if all_extracted_rows:
    df = pd.DataFrame(all_extracted_rows)
    os.makedirs("output_schemas", exist_ok=True)
    output_path = "output_schemas/all_26k_companies_tor.xlsx"
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

            val = str(cell.value or "")
            for line in val.split("\n"):
                if len(line) > max_length:
                    max_length = len(line)

        adjusted_width = max(max_length + 4, 22)
        if adjusted_width > 55:
            adjusted_width = 55
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 75

    wb.save(output_path)
    print(f"[+] SUCCESS! Extracted {len(df)} unique records into {output_path}")
else:
    print("[!] No records extracted.")