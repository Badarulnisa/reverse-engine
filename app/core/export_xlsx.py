import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def csv_to_formatted_xlsx(csv_path: str, xlsx_path: str, sheet_name: str = "Registry") -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        wb.save(xlsx_path)
        return 0

    header, data_rows = rows[0], rows[1:]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(header)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in data_rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows)}"

    for col_idx, col_name in enumerate(header, start=1):
        max_len = len(col_name)
        for row in data_rows:
            if col_idx - 1 < len(row) and row[col_idx - 1]:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(xlsx_path)
    return len(data_rows)

