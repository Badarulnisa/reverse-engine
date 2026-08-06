"""
Builds the final .xlsx report from extracted data. Design goals driven
directly by the spec:
  - nothing missing: every column that appears in ANY row of a table is
    included as a header (union of keys), missing values just render blank
  - spacious, non-overlapping: styled header row, autofit column widths
    (capped so nothing runs absurdly wide), frozen header row, real
    Excel Tables (not raw ranges) so columns/rows never visually collide
  - one Overview sheet + one sheet per detected system/endpoint pair
"""

import re
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")


def _unique(name: str, used: set, max_len: int) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "_", name).strip("_ ") or "sheet"
    clean = clean[:max_len]
    base, i = clean, 1
    while clean.lower() in used:
        suffix = f"_{i}"
        clean = base[: max_len - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def _write_table(ws, rows: List[Dict[str, Any]], table_name: str):
    if not rows:
        return

    # union of every key across every row -- this is what guarantees
    # nothing gets silently dropped from the report
    columns: List[str] = []
    seen = set()
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)

    ws.append(columns)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(c, "") for c in columns])

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).font = BODY_FONT

    # autofit-ish column widths, capped so long JSON blobs don't wreck layout
    for i, col in enumerate(columns, start=1):
        max_len = len(str(col))
        for row in rows:
            max_len = max(max_len, len(str(row.get(col, ""))))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    last_col = get_column_letter(len(columns))
    table = Table(displayName=table_name, ref=f"A1:{last_col}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(table)


def build_report(
    summaries: List[Dict[str, Any]],
    system_endpoint_data: Dict[str, Dict[str, List[Dict[str, Any]]]],
    output_path: str,
    findings: List[Dict[str, Any]] = None,
    finding_summary: List[Dict[str, Any]] = None,
) -> str:
    """
    summaries: output of system_detector.summarize_systems()
    system_endpoint_data: {system_host: {endpoint_path: [flat_rows]}}
    findings: output of vuln_scanner.scan_records() -- optional
    finding_summary: output of vuln_scanner.summarize_findings() -- optional
    """
    wb = Workbook()
    used_sheet_names: set = set()
    used_table_names: set = set()

    overview = wb.active
    overview.title = "Overview"
    used_sheet_names.add("overview")
    overview_table_name = _unique("Overview", used_table_names, 60)
    _write_table(overview, summaries, overview_table_name)

    if finding_summary:
        sec_summary_ws = wb.create_sheet(_unique("Security Summary", used_sheet_names, 31))
        _write_table(sec_summary_ws, finding_summary, _unique("SecuritySummary", used_table_names, 60))

    if findings:
        sec_ws = wb.create_sheet(_unique("Security Findings", used_sheet_names, 31))
        _write_table(sec_ws, findings, _unique("SecurityFindings", used_table_names, 60))

    for system, endpoints in system_endpoint_data.items():
        for endpoint, rows in endpoints.items():
            if not rows:
                continue
            sheet_name = _unique(f"{system}_{endpoint}", used_sheet_names, 31)
            table_name = _unique(f"tbl_{sheet_name}", used_table_names, 60)
            ws = wb.create_sheet(sheet_name)
            _write_table(ws, rows, table_name)

    wb.save(output_path)
    return output_path
