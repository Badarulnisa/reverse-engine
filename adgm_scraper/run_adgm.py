"""
Driver: partitioned crawl of the ADGM public register.

FIXES applied in this version (see prior review for full bug list):
1. listing_complete now correctly requires ALL of a split status's
   category sub-buckets to be done, not just "at least one" (was using
   any(), silently treating partial coverage as complete).
2. A split status's own top-level key IS now added to completed_buckets
   once all its children finish, so resuming doesn't needlessly re-probe
   and re-split it every run.
3. A THIRD split level (name-prefix, a-z0-9) is now implemented for any
   status x category bucket still over the ceiling, instead of just
   logging an error and silently dropping that data.
4. fmt_person now extracts just the clean display name instead of a raw
   key=value dump -- matches the clean DIFC-style export look.

Usage:
    python run_adgm.py --probe
    python run_adgm.py --probe-buckets
    python run_adgm.py
    python run_adgm.py --no-details
    python run_adgm.py --limit 50 --workers 1
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import stat
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from adgm_scraper import (
    AdgmScraper, AuraApiError, StaleContextError,
    ENTITY_STATUS_VALUES, CATEGORY_VALUES, OFFSET_CEILING_ROWS,
)

try:
    from monitor import start_monitor, log_companies_found
    _MONITOR_AVAILABLE = True
except ImportError:
    _MONITOR_AVAILABLE = False
    def log_companies_found(*a, **k): pass

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("run_adgm")

CHECKPOINT_PATH = "adgm_checkpoint.json"
OUTPUT_XLSX = "adgm_companies.xlsx"
SAFE_BUCKET_LIMIT = 1900
ALPHABET = "abcdefghijklmnopqrstuvwxyz0123456789"

company_rows: dict[str, dict] = {}
detail_done_ids: set[str] = set()
completed_buckets: set[str] = set()  # "status", "status||category", or "status||category||prefix"
bucket_progress: dict[str, int] = {}  # bucket_key -> last successfully completed page number
listing_complete: bool = False

_checkpoint_lock = Lock()

_GENERIC_SKIP_FIELDS = {"Id", "RecordTypeId", "attributes", "LastModifiedDate", "_objectName", "_bucketStatus"}
_PERSON_LABELS = {
    "Shareholder", "Director", "Secretary", "Beneficial_Owners",
    "Partners", "Non_Cell_Members", "Cell_Members", "Members",
    "Authorised_Signatories", "Parent_Company_Members", "Parent_Company_Shareholder",
}
_FILING_LABELS = {"Filings"}
_ADDRESS_LABELS = {"Addresses"}

# Fields that hold a person/entity's display name, in priority order --
# used by fmt_person() to show a clean name instead of a raw field dump.
_NAME_FIELD_CANDIDATES = ["Role_Full_Name__c", "Contact_Full_Name__c", "Entity_Name_t__c", "Name"]
_DATE_FIELD_CANDIDATES = ["Appointment_Date__c"]
_STATUS_FIELD = "_bucketStatus"


# ---------------------------------------------------------------------------
# checkpoint
# ---------------------------------------------------------------------------

def load_checkpoint():
    global company_rows, detail_done_ids, completed_buckets, bucket_progress, listing_complete
    if not os.path.exists(CHECKPOINT_PATH):
        return
    with open(CHECKPOINT_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    company_rows = data.get("company_rows", {})
    detail_done_ids = set(data.get("detail_done_ids", []))
    completed_buckets = set(data.get("completed_buckets", []))
    bucket_progress = data.get("bucket_progress", {})
    listing_complete = data.get("listing_complete", False)
    log.info("resumed checkpoint: %d companies, %d buckets done (complete=%s), %d details done, "
             "%d buckets with partial page progress",
             len(company_rows), len(completed_buckets), listing_complete, len(detail_done_ids),
             len(bucket_progress))


def save_checkpoint():
    tmp = CHECKPOINT_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump({
            "company_rows": company_rows,
            "detail_done_ids": list(detail_done_ids),
            "completed_buckets": list(completed_buckets),
            "bucket_progress": bucket_progress,
            "listing_complete": listing_complete,
        }, f, ensure_ascii=False)

    if os.path.exists(CHECKPOINT_PATH):
        try:
            os.chmod(CHECKPOINT_PATH, stat.S_IWRITE)
        except OSError:
            pass

    for attempt in range(6):
        try:
            os.replace(tmp, CHECKPOINT_PATH)
            return
        except PermissionError:
            time.sleep(0.3 * (attempt + 1))

    # last resort: non-atomic overwrite so a stubborn transient lock
    # doesn't lose this checkpoint entirely
    log.warning("save_checkpoint: os.replace failed 6x, falling back to copyfile")
    shutil.copyfile(tmp, CHECKPOINT_PATH)
    os.remove(tmp)


# ---------------------------------------------------------------------------
# listing pass -- partitioned by Entity_Status__c, then Category__c,
# then (FIX #3) name-prefix as a third fallback level
# ---------------------------------------------------------------------------

def flatten_listing_row(row: dict) -> dict:
    addr = ""
    addrs = row.get("Addresses__r") or []
    if addrs:
        addr = addrs[0].get("Full_Address__c") or addrs[0].get("Address_for_DDP__c") or ""

    trade_name = ""
    trades = row.get("Trade_Names__r") or []
    if trades:
        trade_name = trades[0].get("Name_in_English__c", "")

    return {
        "Id": row.get("Id"),
        "Name": row.get("Name"),
        "Trade_Name": trade_name,
        "Entity_Type": row.get("Entity_Type__c"),
        "Entity_Sub_Type": row.get("Entity_Sub_Type__c"),
        "Category": row.get("Category__c"),
        "Entity_Status": row.get("Entity_Status__c"),
        "Registration_Number": row.get("Registration_Number__c"),
        "Incorporation_Date": row.get("Incorporation_Date__c"),
        "Is_Continued": row.get("Is_continued__c"),
        "Address": addr,
    }


def _paginate_bucket(scraper: AdgmScraper, bucket_key: str, entity_status: str,
                      category: str, name_prefix: str = ""):
    """
    Pages one status[/category[/prefix]] bucket to completion, or until it
    hits the offset ceiling (meaning it needs splitting further).

    Resumes from bucket_progress[bucket_key] instead of always starting at
    page 1 -- previously a mid-bucket failure (e.g. a network drop at page
    163) meant every retry re-fetched everything from page 1 again. Since
    company_rows dedupes by Id this never corrupted data, but on a large
    bucket (up to ~1990 rows / ~199 pages) it wasted a lot of time
    re-requesting pages already known to be done. Progress is saved after
    every page, not just at bucket completion, so even a page-162 failure
    loses at most one page of work, not the whole bucket.
    """
    start_page = bucket_progress.get(bucket_key, 0) + 1
    page_number = start_page
    added = 0
    seen_this_bucket = (start_page - 1) * 10  # approx rows already accounted for from prior pages

    if start_page > 1:
        log.info("bucket %s: resuming from page %d (already completed pages 1-%d)",
                 bucket_key, start_page, start_page - 1)

    while True:
        try:
            rows, requestcount = scraper.search_page_with_count(
                name_prefix, page_number, page_size=10, entity_status=entity_status, category=category)
        except StaleContextError:
            raise
        except AuraApiError as e:
            log.error("bucket %s page %d: FAILED (%s) -- will retry from this exact page next run",
                      bucket_key, page_number, e)
            return added, False

        if not rows:
            break

        for row in rows:
            rid = row.get("Id")
            if rid and rid not in company_rows:
                company_rows[rid] = flatten_listing_row(row)
                added += 1
        seen_this_bucket += len(rows)

        # record this page as done BEFORE deciding whether to continue --
        # if the bucket finishes normally or hits the ceiling, this page's
        # data is safely counted either way.
        bucket_progress[bucket_key] = page_number
        save_checkpoint()

        if len(rows) < 10:
            break
        if seen_this_bucket >= OFFSET_CEILING_ROWS - 10:
            log.warning("bucket %s hit the offset ceiling (~%d rows) before finishing "
                        "(requestcount=%s) -- needs splitting further",
                        bucket_key, seen_this_bucket, requestcount)
            return added, "too_big"

        page_number += 1
        time.sleep(scraper.request_delay)

    return added, True


def _crawl_status_category(scraper: AdgmScraper, status: str, category: str) -> bool:
    """
    Handles one status x category combo. If it fits, paginate it directly.
    If it's still over the ceiling, split by name-prefix (a-z0-9) as a
    third level. Returns True only if this bucket (and all its children,
    if split further) is genuinely, fully complete -- the caller uses this
    to decide whether it's safe to mark the parent status as done. A
    network failure partway through must NOT be reported as success, or
    the parent status gets falsely marked complete and the remaining data
    is silently skipped on every future resume.
    """
    sub_key = f"{status}||{category}"
    if sub_key in completed_buckets:
        return True

    added, result = _paginate_bucket(scraper, sub_key, entity_status=status, category=category)

    if result is True:
        completed_buckets.add(sub_key)
        log.info("  bucket %r done: +%d companies", sub_key, added)
        log_companies_found(len(company_rows))
        save_checkpoint()
        return True

    if result == "too_big":
        log.info("  bucket %r still too big -- splitting by name prefix (a-z0-9)...", sub_key)
        all_prefixes_ok = True
        for ch in ALPHABET:
            prefix_key = f"{status}||{category}||{ch}"
            if prefix_key in completed_buckets:
                continue
            p_added, p_result = _paginate_bucket(scraper, prefix_key, entity_status=status,
                                                  category=category, name_prefix=ch)
            if p_result is True:
                completed_buckets.add(prefix_key)
                log_companies_found(len(company_rows))
            elif p_result == "too_big":
                log.error("  prefix bucket %r STILL over the ceiling even after 3 split levels -- "
                          "this is a genuinely dense segment. Consider a 4th split (two-letter "
                          "prefix) manually if this data matters. Skipping for now.", prefix_key)
                all_prefixes_ok = False
            else:
                # failed (network error etc) -- not complete, will retry
                all_prefixes_ok = False
            save_checkpoint()

        if all_prefixes_ok:
            completed_buckets.add(sub_key)
            save_checkpoint()
            return True
        else:
            log.warning("  bucket %r NOT fully complete (some prefix sub-buckets failed or are "
                        "still too big) -- will retry remaining ones next run", sub_key)
            return False

    # result is False (failed, e.g. network error) -- definitely not complete
    log.warning("  bucket %r failed to complete -- will retry next run", sub_key)
    return False


def run_full_listing_pass(scraper: AdgmScraper):
    global listing_complete
    if listing_complete:
        log.info("Listing already complete (%d companies known)", len(company_rows))
        return

    for status in ENTITY_STATUS_VALUES:
        # FIX #2: check the status's OWN key first -- if it's already
        # marked done (whether resolved directly or after a full split),
        # skip re-probing it entirely on resume.
        if status in completed_buckets:
            continue

        log.info("=== Bucket: Entity_Status__c=%r ===", status)
        added, result = _paginate_bucket(scraper, status, entity_status=status, category="")

        if result is True:
            completed_buckets.add(status)
            log.info("bucket %r done: +%d companies (total so far: %d)", status, added, len(company_rows))
            log_companies_found(len(company_rows))
            save_checkpoint()
        elif result == "too_big":
            log.info("splitting bucket %r by Category__c...", status)
            # CRITICAL FIX: only mark the parent status as done if EVERY
            # category sub-bucket genuinely completed. Previously this
            # unconditionally marked the status done after the loop, even
            # if a category errored out partway (e.g. a network drop) --
            # that silently truncated ~14,000-company statuses like
            # "Registered" and made the listing pass falsely report
            # complete with only a few thousand companies collected.
            all_categories_ok = True
            for category in CATEGORY_VALUES:
                ok = _crawl_status_category(scraper, status, category)
                if not ok:
                    all_categories_ok = False
            if all_categories_ok:
                completed_buckets.add(status)
                log.info("bucket %r fully done across all categories (total so far: %d)",
                         status, len(company_rows))
            else:
                log.warning("bucket %r NOT fully done -- some categories failed/incomplete, "
                            "will retry remaining ones next run", status)
            save_checkpoint()
        else:
            log.warning("bucket %r incomplete, will retry next run", status)

    # FIX #1: listing is complete only when EVERY status value's own key
    # is in completed_buckets -- no longer accepts partial category
    # coverage as "good enough".
    listing_complete = all(status in completed_buckets for status in ENTITY_STATUS_VALUES)
    save_checkpoint()
    log.info("Listing pass complete: %d unique companies known across %d buckets (full=%s).",
             len(company_rows), len(completed_buckets), listing_complete)


# ---------------------------------------------------------------------------
# detail pass
# ---------------------------------------------------------------------------

def fmt_person(rec: dict) -> str:
    """FIX #4: clean display instead of a raw key=value dump. Shows the
    name, status, and appointment date only -- the fields a human actually
    wants to see, matching the DIFC-style clean export look."""
    name = ""
    for field_name in _NAME_FIELD_CANDIDATES:
        if rec.get(field_name):
            name = rec[field_name]
            break
    if not name:
        return ""
    status = rec.get(_STATUS_FIELD, "")
    date = ""
    for field_name in _DATE_FIELD_CANDIDATES:
        if rec.get(field_name):
            date = rec[field_name]
            break
    parts = [name]
    if status:
        parts.append(f"({status})")
    if date:
        parts.append(f"since {date}")
    return " ".join(parts)


def fmt_address(rec: dict) -> str:
    """Clean single-line address instead of a key=value dump."""
    addr = rec.get("Full_Address__c") or rec.get("Address_for_DDP__c") or ""
    return addr


def fmt_filing(rec: dict) -> str:
    parts = []
    for k, v in rec.items():
        if k in _GENERIC_SKIP_FIELDS or v in (None, "", []):
            continue
        parts.append(f"{k}={v}")
    return "; ".join(parts)


def fmt_generic(rec: dict) -> str:
    rtype = rec.get("_recordTypeName") or (rec.get("RecordType") or {}).get("DeveloperName")
    prefix = f"[{rtype}] " if rtype else ""
    parts = []
    for k, v in rec.items():
        if k in _GENERIC_SKIP_FIELDS or v in (None, "", []):
            continue
        if isinstance(v, dict):
            sub = ", ".join(f"{sk}={sv}" for sk, sv in v.items() if sv not in (None, ""))
            if sub:
                parts.append(f"{k}=({sub})")
            continue
        parts.append(f"{k}={v}")
    return prefix + "; ".join(parts)


def extract_records_from_entry(entry: dict) -> list[dict]:
    out: list[dict] = []
    seen_ids: set[str] = set()
    for bucket in entry.get("records", []) or []:
        object_name = bucket.get("objectName")
        for group in bucket.get("activeInactiveRecords", []) or []:
            record_type_name = group.get("recordTypeName")
            status = group.get("Status")
            for rec in group.get("records", []) or []:
                if not isinstance(rec, dict):
                    continue
                rid = rec.get("Id")
                if rid and rid in seen_ids:
                    continue
                if rid:
                    seen_ids.add(rid)
                tagged = dict(rec)
                tagged.setdefault("_objectName", object_name)
                tagged.setdefault("_recordTypeName", record_type_name)
                tagged.setdefault("_bucketStatus", status)
                out.append(tagged)
    return out


def extract_business_activity_type(all_pages: dict) -> str:
    entry = (all_pages or {}).get("Business_Activities")
    if not entry:
        return ""
    names = []
    for rec in extract_records_from_entry(entry):
        name = rec.get("Activity_Name__c")
        if name and name not in names:
            names.append(name)
    return "; ".join(names)


def flatten_detail_records(all_pages: dict) -> dict:
    out = {}
    for label, entry in (all_pages or {}).items():
        records = extract_records_from_entry(entry)
        if not records:
            continue

        formatter = (fmt_person if label in _PERSON_LABELS
                     else fmt_filing if label in _FILING_LABELS
                     else fmt_address if label in _ADDRESS_LABELS
                     else fmt_generic)

        groups: dict[str, list[dict]] = {}
        for rec in records:
            rtype = rec.get("_recordTypeName") or ""
            groups.setdefault(rtype, []).append(rec)

        for rtype, recs in groups.items():
            col_name = f"{label}__{rtype}" if rtype else label
            formatted = [formatter(r) for r in recs]
            formatted = [f for f in formatted if f]  # drop empties
            out[col_name] = " | ".join(formatted)

    out["Business_Activity_Type"] = extract_business_activity_type(all_pages)
    return out


def _fetch_one_detail(scraper: AdgmScraper, cid: str):
    try:
        all_pages = scraper.get_all_detail_pages(cid)
        return cid, flatten_detail_records(all_pages), None
    except StaleContextError:
        raise
    except AuraApiError as e:
        return cid, None, e


def process_details(scraper: AdgmScraper, limit: int | None = None, max_workers: int = 3):
    ids_to_process = [cid for cid in company_rows if cid not in detail_done_ids]
    if limit:
        ids_to_process = ids_to_process[:limit]
    log.info("Detail pass: %d companies pending (max_workers=%d)", len(ids_to_process), max_workers)
    if not ids_to_process:
        return

    def make_worker_scraper(i):
        s = AdgmScraper(fwuid=scraper.fwuid, app=scraper.app, loaded=dict(scraper.loaded),
                         worker_name=f"w{i}", use_tor=scraper.use_tor)
        s.bootstrap()
        return s

    worker_scrapers = [make_worker_scraper(i) for i in range(max_workers)]

    done_count = 0
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {}
        for i, cid in enumerate(ids_to_process):
            ws = worker_scrapers[i % max_workers]
            fut = pool.submit(_fetch_one_detail, ws, cid)
            futures[fut] = cid

        for fut in as_completed(futures):
            cid = futures[fut]
            row = company_rows[cid]
            try:
                cid, flat, err = fut.result()
            except StaleContextError as e:
                log.error("STALE CONTEXT -- stopping. Progress saved, refresh fwuid in adgm_scraper.py. %s", e)
                with _checkpoint_lock:
                    save_checkpoint()
                sys.exit(1)

            if err is not None:
                log.error("detail fetch failed for %s (%s): %s -- will retry next run", cid, row.get("Name"), err)
            else:
                with _checkpoint_lock:
                    row.update(flat)
                    detail_done_ids.add(cid)

            done_count += 1
            log.info("[%d/%d] done: %s", done_count, len(ids_to_process), row.get("Name"))
            if done_count % 20 == 0:
                with _checkpoint_lock:
                    save_checkpoint()

    with _checkpoint_lock:
        save_checkpoint()


# ---------------------------------------------------------------------------
# export -- clean, separated columns
# ---------------------------------------------------------------------------

def _ordered_columns(all_keys: list[str]) -> list[str]:
    priority = [
        "Id", "Name", "Trade_Name", "Entity_Type", "Entity_Sub_Type",
        "Category", "Entity_Status", "Registration_Number",
        "Incorporation_Date", "Is_Continued", "Address", "Business_Activity_Type",
    ]
    person_prefixes = tuple(_PERSON_LABELS)
    ordered = [k for k in priority if k in all_keys]
    remaining = [k for k in all_keys if k not in ordered]
    person_cols = sorted(k for k in remaining if k.startswith(person_prefixes))
    other_cols = sorted(k for k in remaining if not k.startswith(person_prefixes))
    return ordered + person_cols + other_cols


def export_xlsx():
    if not company_rows:
        log.warning("nothing to export")
        return

    all_keys_unordered: list[str] = []
    seen = set()
    for row in company_rows.values():
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                all_keys_unordered.append(k)
    all_keys = _ordered_columns(all_keys_unordered)

    wb = Workbook()
    ws = wb.active
    ws.title = "ADGM Companies"
    ws.append(all_keys)
    for row in company_rows.values():
        ws.append([row.get(k, "") for k in all_keys])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                          top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
            val = str(cell.value or "")
            for line in val.split("\n"):
                max_length = max(max_length, len(line))
        ws.column_dimensions[col_letter].width = min(max(max_length + 4, 22), 55)
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60

    wb.save(OUTPUT_XLSX)
    log.info("Exported %d companies to %s", len(company_rows), OUTPUT_XLSX)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    if _MONITOR_AVAILABLE:
        start_monitor()

    parser = argparse.ArgumentParser()
    parser.add_argument("--no-details", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--workers", type=int, default=3)
    parser.add_argument("--resume-only-report", action="store_true")
    parser.add_argument("--probe", action="store_true", help="Report total requestcount, exit")
    parser.add_argument("--probe-buckets", action="store_true",
                         help="Report requestcount per Entity_Status__c value, exit")
    parser.add_argument("--tor", action="store_true",
                         help="Route requests through Tor (curl_cffi + SOCKS5), rotating circuit "
                              "on connection errors / 403 / 429 / 503. Requires Tor running "
                              "locally and TOR_CONTROL_PASSWORD set. See tor_rotation.py.")
    args = parser.parse_args()

    if args.tor:
        # This project's torrc uses CookieAuthentication (no password
        # configured), so TOR_CONTROL_PASSWORD is optional -- only warn,
        # don't require it. See tor_rotation.py.
        import os
        if not os.environ.get("TOR_CONTROL_PASSWORD"):
            log.info(
                "--tor: TOR_CONTROL_PASSWORD not set -- will use Tor's cookie "
                "authentication instead (torrc has CookieAuthentication 1). "
                "Make sure Tor is running (tor.exe -f torrc) before continuing."
            )

    if args.probe:
        scraper = AdgmScraper(use_tor=args.tor)
        try:
            rows, requestcount = scraper.search_page_with_count("", 1, page_size=10)
        except (StaleContextError, AuraApiError) as e:
            log.error("Probe failed: %s", e)
            sys.exit(1)
        print(f"Total entities on the register: {requestcount}")
        print(f"That needs {-(-requestcount // 10)} pages of 10 to fully paginate.")
        print(f"Offset ceiling means flat pagination alone stalls around row {OFFSET_CEILING_ROWS} -- "
              f"use --probe-buckets to check the partitioned approach.")
        return

    if args.probe_buckets:
        scraper = AdgmScraper(use_tor=args.tor)
        print("=== Probing Entity_Status__c buckets (confirmed real values) ===")
        total = 0
        for status in ENTITY_STATUS_VALUES:
            rc = scraper.probe_requestcount(entity_status=status)
            flag = ""
            if rc is not None:
                total += rc
                if rc > SAFE_BUCKET_LIMIT:
                    flag = f"  <-- TOO BIG (>{SAFE_BUCKET_LIMIT}), will auto-split by Category__c (then name-prefix) at crawl time"
            print(f"  {status!r:30} -> requestcount={rc}{flag}")
            time.sleep(scraper.request_delay)
        print(f"\nSum across all status buckets: {total} (compare to unfiltered total via --probe)")
        return

    load_checkpoint()

    if args.resume_only_report:
        print(f"Companies known: {len(company_rows)}")
        print(f"Buckets done: {len(completed_buckets)} (complete={listing_complete})")
        print(f"Detail-fetches completed: {len(detail_done_ids)}")
        return

    scraper = AdgmScraper(use_tor=args.tor)
    try:
        log.info("=== LISTING PASS (partitioned) ===")
        run_full_listing_pass(scraper)

        if not args.no_details:
            log.info("=== DETAIL PASS ===")
            process_details(scraper, limit=args.limit, max_workers=args.workers)

    except StaleContextError as e:
        log.error("STALE CONTEXT -- stopping. Progress saved, refresh fwuid/loaded in adgm_scraper.py. %s", e)
        save_checkpoint()
        sys.exit(1)
    except KeyboardInterrupt:
        log.warning("Interrupted -- progress saved, re-run to resume.")
        save_checkpoint()
        sys.exit(1)

    export_xlsx()


if __name__ == "__main__":
    main()