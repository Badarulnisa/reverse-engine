import os
import time
import json
import random
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

URL = "https://www.difc.com/api/handleRequest"
CHECKPOINT_PATH = "difc_checkpoint.json"
OUTPUT_PATH = "output_schemas/all_difc_companies.xlsx"

PAGE_SIZE = 10          # confirmed fixed page size from HAR
MAX_WORKERS = 6          # concurrent requests — polite but no longer sequential
REQUEST_DELAY = (0.2, 0.5)   # random jitter range per request, seconds

HEADERS = {
    "accept": "*/*",
    "content-type": "text/plain;charset=UTF-8",
    "origin": "https://www.difc.com",
    "referer": "https://www.difc.com/business/public-register",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
}

session = requests.Session()
session.headers.update(HEADERS)

all_rows_by_id = {}
completed_queries = set()   # set of name-query strings we've fully paged through
completed_license_nos = set()   # set of license numbers already checked in the numeric sweep
exhausted = False   # kept for backward-compat with old checkpoints; no longer drives stopping


def load_checkpoint():
    global all_rows_by_id, completed_queries, exhausted, completed_license_nos
    if os.path.exists(CHECKPOINT_PATH):
        with open(CHECKPOINT_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        all_rows_by_id = data.get("rows", {})
        completed_queries = set(data.get("completed_queries", []))
        completed_license_nos = set(data.get("completed_license_nos", []))
        exhausted = data.get("exhausted", False)
        print(f"[i] Resumed: {len(all_rows_by_id)} rows, {len(completed_queries)} name-queries done, "
              f"{len(completed_license_nos)} license-numbers checked")


def save_checkpoint():
    tmp_path = CHECKPOINT_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump({
            "rows": all_rows_by_id,
            "completed_queries": sorted(completed_queries),
            "completed_license_nos": sorted(completed_license_nos),
            "exhausted": exhausted,
        }, f, ensure_ascii=False)
    os.replace(tmp_path, CHECKPOINT_PATH)


def fetch_query_page(name_query: str, offset: int):
    """
    Single API call for one (name_query, offset) page. Returns (rows, success).
    Caller is responsible for detecting duplicate/repeat pages — this endpoint
    does NOT reliably return an empty list at the true end of results; past
    some point it silently loops and re-serves already-seen records instead.
    """
    payload = {
        "name": name_query,
        "licenseType": "",
        "licenseNo": "",
        "status": "",
        "offset": offset,
        "slug": "/CRM/public-register",
        "method": "POST",
    }
    try:
        response = session.post(URL, data=json.dumps(payload), timeout=30)
        if response.status_code != 200:
            print(f"    [!] HTTP {response.status_code} on name='{name_query}' offset={offset}")
            return [], False
        parsed = response.json()
        if not parsed.get("IsSuccess", True):
            print(f"    [!] API error on name='{name_query}' offset={offset}: {parsed.get('Message')}")
            return [], False
        company_list = parsed.get("Data", {}).get("companyList", []) or []
        return company_list, True
    except Exception as e:
        print(f"    [!] Exception on name='{name_query}' offset={offset}: {e}")
        return [], False


def scrape_query(name_query: str, max_offset: int = 5000):
    """
    Pages through a single name-filter query sequentially, stopping as soon as
    a page returns ZERO new (not-yet-seen) company Ids — this catches both the
    honest-empty-page case AND the silent-repeat-loop case in one check, since
    a looping page will consist entirely of Ids we've already recorded.
    """
    offset = 0
    added_this_query = 0
    consecutive_failures = 0

    while offset <= max_offset:
        rows, success = fetch_query_page(name_query, offset)
        time.sleep(random.uniform(*REQUEST_DELAY))

        if not success:
            consecutive_failures += 1
            if consecutive_failures >= 3:
                print(f"    [!] '{name_query}': 3 consecutive failures, aborting this query (will retry next run)")
                return added_this_query, False
            continue
        consecutive_failures = 0

        if not rows:
            break  # honest empty page

        new_ids_this_page = 0
        for row in rows:
            rid = row.get("Id") or f"noId_{row.get('Registration_License_No__c')}"
            if rid not in all_rows_by_id:
                all_rows_by_id[rid] = normalize_row(row)
                added_this_query += 1
                new_ids_this_page += 1

        if new_ids_this_page == 0:
            # every record on this page was already known -> repeat/loop detected, real end reached
            break

        offset += PAGE_SIZE

    return added_this_query, True


def generate_queries():
    """
    Two-letter lowercase prefixes (aa, ab, ..., zz) plus single digits, since
    single letters alone were confirmed to loop past ~130 unique matches.
    676 letter combos + 10 digit combos = 686 queries total.
    """
    import string
    letters = string.ascii_lowercase
    for a in letters:
        for b in letters:
            yield a + b
    for d in string.digits:
        yield d


LICENSE_SWEEP_START = 1
LICENSE_SWEEP_END = 14500  # confirmed real companies exist up to ~14200; small buffer added


def fetch_license_no(license_no: int):
    """
    Exact-match lookup for a single license number. Confirmed live: full
    numbers (e.g. "9050") match exactly one company; only short/partial
    numbers behave as substring matches, which is why the sweep always
    sends the complete integer, never a prefix.
    """
    payload = {
        "name": "",
        "licenseType": "",
        "licenseNo": str(license_no),
        "status": "",
        "offset": 0,
        "slug": "/CRM/public-register",
        "method": "POST",
    }
    try:
        response = session.post(URL, data=json.dumps(payload), timeout=30)
        if response.status_code != 200:
            return license_no, [], False
        parsed = response.json()
        if not parsed.get("IsSuccess", True):
            return license_no, [], False
        company_list = parsed.get("Data", {}).get("companyList", []) or []
        # keep only the exact match, in case of any substring bleed-through
        exact = [c for c in company_list if c.get("Registration_License_No__c") == str(license_no)]
        return license_no, exact, True
    except Exception as e:
        print(f"    [!] Exception on licenseNo={license_no}: {e}")
        return license_no, [], False


def sweep_license_numbers():
    """
    Numeric sweep to catch companies the name-prefix search structurally
    cannot reach (confirmed real: e.g. license 12762 existed but was never
    surfaced by any of the 686 name queries). Runs concurrently since each
    lookup is independent — no shared offset/dedup state like scrape_query.
    """
    pending = [n for n in range(LICENSE_SWEEP_START, LICENSE_SWEEP_END + 1)
               if n not in completed_license_nos]
    total = len(pending)
    if total == 0:
        print("[i] License-number sweep already complete.")
        return

    print(f"[i] {total} license numbers left to check (range {LICENSE_SWEEP_START}-{LICENSE_SWEEP_END}).")

    done_count = 0
    added_count = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(fetch_license_no, n): n for n in pending}
        for fut in as_completed(futures):
            time.sleep(random.uniform(*REQUEST_DELAY))
            license_no, rows, success = fut.result()

            if success:
                completed_license_nos.add(license_no)
                for row in rows:
                    rid = row.get("Id") or f"noId_{row.get('Registration_License_No__c')}"
                    if rid not in all_rows_by_id:
                        all_rows_by_id[rid] = normalize_row(row)
                        added_count += 1
            # if not success, leave it out of completed_license_nos so it retries next run

            done_count += 1
            if done_count % 200 == 0 or done_count == total:
                print(f"    [+] License sweep progress: {done_count}/{total} checked, "
                      f"+{added_count} new companies found (running total unique: {len(all_rows_by_id)})")
                save_checkpoint()

    save_checkpoint()
    print(f"[*] License sweep complete: {done_count}/{total} checked, +{added_count} new companies found.")


def run():
    load_checkpoint()

    queries = [q for q in generate_queries() if q not in completed_queries]
    total_queries = len(queries) + len(completed_queries)

    if queries:
        print(f"[i] {len(queries)} of {total_queries} name-prefix queries remaining.")
        for i, q in enumerate(queries, 1):
            added, ok = scrape_query(q)
            if ok:
                completed_queries.add(q)
                print(f"[+] '{q}' ({i}/{len(queries)}): +{added} new "
                      f"(running total unique: {len(all_rows_by_id)})")
            else:
                print(f"[i] '{q}' will be retried on next run.")

            if i % 10 == 0:
                save_checkpoint()

        save_checkpoint()
        print(f"\n[*] Name-prefix listing pass complete. Total unique records: {len(all_rows_by_id)}")
    else:
        print("[i] All name-prefix queries already completed.")

    print("\n[*] Starting license-number sweep — catches companies name-search structurally misses...")
    sweep_license_numbers()

    print("\n[*] Starting detail pass — fetching full record per company...")
    fetch_details_for_all()

    print(f"\n[*] Extraction complete. Total unique records: {len(all_rows_by_id)}")
    export()


def fetch_detail(company_id: str):
    """
    Detail-page call for a single company. Confirmed live from the site's own
    network traffic: the backend proxy expects the recordId embedded in the
    slug's querystring, with method set to GET (not POST like the listing call).
    """
    payload = {
        "slug": f"/CRM/public-register?recordId={company_id}",
        "method": "GET",
    }
    try:
        response = session.post(URL, data=json.dumps(payload), timeout=30)
        if response.status_code != 200:
            return company_id, None, False
        parsed = response.json()
        if not parsed.get("IsSuccess", True):
            return company_id, None, False
        records = parsed.get("Data", {}).get("DIFCData", {}).get("PublicRegistry", [])
        if not records:
            return company_id, None, False
        return company_id, records[0], True
    except Exception as e:
        print(f"    [!] Exception fetching detail for {company_id}: {e}")
        return company_id, None, False


def names_joined(entries, name_key, active_only=True):
    """Collapse a list of {NameKey, ConfirmCessation, ...} dicts into a semicolon string."""
    if not entries:
        return ""
    out = []
    for e in entries:
        if active_only and str(e.get("ConfirmCessation", "")).lower() not in ("not ceased", ""):
            continue
        val = e.get(name_key)
        if val:
            out.append(val)
    return "; ".join(out)


def normalize_detail(record: dict) -> dict:
    """Map DIFC's rich detail-page fields to a clean output schema."""
    return {
        "Registered Number": record.get("RegisteredNumber"),
        "Legal Type of Entity": record.get("LegalTypeOfEntity"),
        "Status": record.get("StatusOfRegistration"),
        "Company Type": record.get("CompanyType"),
        "Type of License": record.get("TypeOfLicense"),
        "Type of Entity": record.get("TypeOfEntity"),
        "Date of Incorporation": record.get("DateOfIncorporation"),
        "License Validity Date": record.get("CommercialLicenseValidityDate"),
        "DNFBP": record.get("DNFBP"),
        "Financial Year End": record.get("FinancialYearEnd"),
        "Data Protection Officer Appointed": record.get("AppointedDataProtectionOfficer"),
        "Entity Name": names_joined(record.get("EntityName"), "Name"),
        "Trading Name": names_joined(record.get("TradingName"), "TradeName"),
        "Licensed Activities": names_joined(record.get("LicencedActivity"), "Activity"),
        "Registered Office Address": names_joined(record.get("RegisteredOfficeAddress"), "OfficeAddress"),
        "Current Directors": names_joined(record.get("Director"), "DirectorName"),
        "Shareholders": names_joined(record.get("ShareHolder"), "ShareholderName"),
        "Company Secretary": names_joined(record.get("CompanySecretary"), "SecretaryName"),
        "Auditor": names_joined(record.get("Auditor"), "NameofAuditorFirm"),
        "Share Capital": "; ".join(
            i.get("AccountShareInfo", "") for i in (record.get("AccountSharesInfo") or []) if i.get("AccountShareInfo")
        ),
    }


def normalize_row(row: dict) -> dict:
    """Map DIFC's raw fields to a clean output schema — no scraper metadata, no confidence scores."""
    return {
        "Company Name": row.get("Name"),
        "Registration/License Number": row.get("Registration_License_No__c"),
        "Status": row.get("ROC_Status__c"),
        "Legal Type": row.get("Legal_Type_of_Entity__c"),
        "Legal Entity Type": row.get("Legal_Entity_Type__c"),
        "Incorporation Date": row.get("ROC_reg_incorp_Date__c"),
        "License Activities": row.get("License_Activity_Details__c"),
        "Nature of Business": row.get("Nature_of_business__c"),
        "Registered Address": row.get("Registered_Address__c"),
        "Company Type": row.get("Company_Type__c"),
        "Website": row.get("Website"),
    }


def fetch_details_for_all():
    """
    Second stage: for every company Id collected in the listing pass, fetch
    its rich detail record (directors, shareholders, secretary, auditor, etc.)
    and merge it into that company's row. Resumable via detail_done flag stored
    per record and periodic checkpoint saves.
    """
    pending_ids = [
        rid for rid, row in all_rows_by_id.items()
        if not row.get("_detail_fetched") and not rid.startswith("noId_")
    ]
    total = len(pending_ids)
    if total == 0:
        print("[i] No pending detail fetches — all records already enriched.")
        return

    print(f"[i] {total} companies need detail data.")

    done_count = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(fetch_detail, rid): rid for rid in pending_ids}
        for fut in as_completed(futures):
            time.sleep(random.uniform(*REQUEST_DELAY))
            company_id, record, success = fut.result()

            if success and record:
                merged = all_rows_by_id.get(company_id, {})
                merged.update(normalize_detail(record))
                merged["_detail_fetched"] = True
                all_rows_by_id[company_id] = merged
            else:
                # leave _detail_fetched unset so it retries on next run
                pass

            done_count += 1
            if done_count % 50 == 0 or done_count == total:
                print(f"    [+] Detail progress: {done_count}/{total}")
                save_checkpoint()

    save_checkpoint()
    print(f"[*] Detail pass complete: {done_count}/{total} processed.")


def export():
    if not all_rows_by_id:
        print("[!] No records extracted.")
        return

    all_extracted_rows = [
        {k: v for k, v in row.items() if k != "_detail_fetched"}
        for row in all_rows_by_id.values()
    ]
    df = pd.DataFrame(all_extracted_rows)
    os.makedirs("output_schemas", exist_ok=True)
    df.to_excel(OUTPUT_PATH, index=False)

    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

            val = str(cell.value or "")
            for line in val.split("\n"):
                if len(line) > max_length:
                    max_length = len(line)

        adjusted_width = min(max(max_length + 4, 22), 55)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 60

    wb.save(OUTPUT_PATH)
    print(f"[+] SUCCESS! Exported {len(df)} unique records into {OUTPUT_PATH}")


if __name__ == "__main__":
    run()