import os
import time
import json
import string
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from app.core.har_parser import parse_har_file

HAR_PATH = "dmcc.ae.har"
CHECKPOINT_PATH = "dmcc_checkpoint.json"
OUTPUT_PATH = "output_schemas/all_26k_companies.xlsx"
CAP = 100  # confirmed hard cap per search term returned by this endpoint

print("[*] Parsing HAR file for session headers, ctx, and endpoint...")
records = parse_har_file(HAR_PATH)

target_rec = None
for rec in records:
    if "apexremote" in rec.get("url", "") and rec.get("method") == "POST":
        target_rec = rec
        break

if not target_rec:
    print("[!] Error: Could not find the Salesforce apexremote POST call in the HAR file.")
    exit(1)

url = target_rec.get("url")
raw_headers = target_rec.get("request_headers", {})
DROP_HEADERS = {"content-length", "host"}
headers = {
    k: v for k, v in raw_headers.items()
    if not k.startswith(":") and k.lower() not in DROP_HEADERS
}
ctx = target_rec.get("request_body", {}).get("ctx")

print(f"[+] Target Endpoint: {url}")
print(f"[+] Loaded {len(headers)} usable headers.")
print(f"[+] ctx present: {bool(ctx)}")

session = requests.Session()
session.headers.update(headers)

all_rows_by_license = {}
completed_terms = set()

def load_checkpoint():
    global all_rows_by_license, completed_terms
    if os.path.exists(CHECKPOINT_PATH):
        with open(CHECKPOINT_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        all_rows_by_license = data.get("rows", {})
        completed_terms = set(data.get("completed_terms", []))
        print(f"[i] Resumed from checkpoint: {len(all_rows_by_license)} rows, {len(completed_terms)} terms done.")

def save_checkpoint():
    tmp_path = CHECKPOINT_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump({
            "rows": all_rows_by_license,
            "completed_terms": sorted(completed_terms),
        }, f, ensure_ascii=False)
    os.replace(tmp_path, CHECKPOINT_PATH)

def fetch_term(term: str):
    """
    Single API call for this exact term. Returns (added_count, success, hit_cap).
    hit_cap=True means exactly CAP results came back -> there are likely more
    real matches hidden behind this term, and it needs to be split further.
    """
    payload = [{
        "action": "DMCCPublicDirectoryPage_Ctrl",
        "method": "searchCustomerDirectory",
        "data": [term, ""],
        "type": "rpc",
        "tid": 1,
        "ctx": ctx
    }]

    try:
        response = session.post(url, json=payload, timeout=60)
        if response.status_code != 200:
            print(f"    [!] HTTP {response.status_code} on term={term!r}: {response.text[:200]}")
            return 0, False, False

        parsed = response.json()
        entry = parsed[0] if isinstance(parsed, list) and parsed else parsed
        status_code = entry.get("statusCode")

        if status_code != 200:
            print(f"    [!] API error on term={term!r}: {entry.get('message')}")
            return 0, False, False

        result_data = entry.get("result", [])
        if not isinstance(result_data, list):
            result_data = []

        added = 0
        for row in result_data:
            lic_no = row.get("licNo")
            key = lic_no or f"noLic_{row.get('customerName')}_{row.get('customerRegistrationNumber')}"
            if key in all_rows_by_license:
                continue

            activities = row.get("licenseActivities", [])
            activities_ar = row.get("licenseArabicActivities", [])

            clean_row = {
                "Company Name (English)": row.get("customerName"),
                "Company Name (Arabic)": row.get("customerArabicName"),
                "License Number": row.get("licNo"),
                "License Issue Date": row.get("licIssueDate"),
                "License Expiry Date": row.get("licExpDate"),
                "License Address (English)": str(row.get("licAddress", "")).replace("<br>", "\n"),
                "License Address (Arabic)": str(row.get("licArabicAddress", "")).replace("<br>", "\n"),
                "License Manager": row.get("licManagerName"),
                "License Manager (Arabic)": row.get("licManagerArabicName"),
                "License Activity": "\n".join(activities) if isinstance(activities, list) else str(activities),
                "License Activity (Arabic)": "\n".join(activities_ar) if isinstance(activities_ar, list) else str(activities_ar),
                "License Status": row.get("licenseStatus"),
                "Registration Status": row.get("customerRegistrationStatus"),
                "Registration Number": row.get("customerRegistrationNumber"),
                "Registration Date": row.get("customerRegistrationDate")
            }
            all_rows_by_license[key] = clean_row
            added += 1

        hit_cap = len(result_data) >= CAP
        return added, True, hit_cap

    except Exception as e:
        print(f"    [!] Exception on term={term!r}: {e}")
        return 0, False, False

def process_term(term: str, depth: int = 0):
    """Fetch a term; if it hit the cap, recursively split into term+letter."""
    if term in completed_terms:
        return

    added, success, hit_cap = fetch_term(term)
    total_now = len(all_rows_by_license)
    indent = "  " * depth

    if not success:
        print(f"{indent}[!] term={term!r}: FAILED, will retry on next run")
        return

    print(f"{indent}[+] term={term!r}: +{added} new (running total unique: {total_now}){' [CAPPED]' if hit_cap else ''}")

    completed_terms.add(term)
    save_checkpoint()
    time.sleep(0.3)

    if hit_cap and depth < 4:  # safety limit on recursion depth
        for letter in string.ascii_lowercase:
            process_term(term + letter, depth + 1)
        # Also try digits, in case company names start with numbers
        for digit in "0123456789":
            process_term(term + digit, depth + 1)

# --- Main run ---
load_checkpoint()

for letter in string.ascii_lowercase:
    process_term(letter)

for digit in "0123456789":
    process_term(digit)

print(f"\n[*] Extraction complete. Total unique records: {len(all_rows_by_license)}")

all_extracted_rows = list(all_rows_by_license.values())
if all_extracted_rows:
    df = pd.DataFrame(all_extracted_rows)
    os.makedirs("output_schemas", exist_ok=True)
    df.to_excel(OUTPUT_PATH, index=False)

    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

            val = str(cell.value or "")
            for line in val.split("\n"):
                if len(line) > max_length:
                    max_length = len(line)

        adjusted_width = min(max(max_length + 4, 22), 55)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 75

    wb.save(OUTPUT_PATH)
    print(f"[+] SUCCESS! Extracted {len(df)} unique records into {OUTPUT_PATH}")
else:
    print("[!] No records extracted.")