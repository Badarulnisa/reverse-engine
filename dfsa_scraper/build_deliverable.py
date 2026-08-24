"""
build_deliverable.py

Reads the internal scraper output (firms.jsonl -- successfully collected
and validated records only; failures already live separately in
errors.jsonl and were never written here) and produces:

  FINAL/DFSA_Companies.xlsx   -- clean, professional, boss-facing workbook
  INTERNAL/                   -- firms.jsonl, errors.jsonl, checkpoint.json,
                                  moved here (not copied) so there is one
                                  unambiguous internal location

Run this any time -- during a paused collection run or after it finishes
-- to regenerate the workbook from whatever has been validated so far.
Safe to re-run repeatedly; it always rebuilds FINAL/DFSA_Companies.xlsx
from scratch rather than appending to a stale one.

Usage:
    python build_deliverable.py
"""
from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SOURCE_FIRMS_JSONL = "firms.jsonl"
SOURCE_ERRORS_JSONL = "errors.jsonl"
SOURCE_CHECKPOINT = "checkpoint.json"

FINAL_DIR = Path("FINAL")
INTERNAL_DIR = Path("INTERNAL")
OUTPUT_XLSX = FINAL_DIR / "DFSA_Companies.xlsx"

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)


def load_valid_records(path: str) -> list[dict]:
    """
    Loads firms.jsonl. Every line in this file already represents a firm
    that completed the full pipeline (Fetch -> Parse -> Validate ->
    Complete) successfully -- failures were routed to errors.jsonl by the
    scraper and never written here. Still re-checks the two fields a
    business-facing sheet cannot be missing (Name, DFSA Reference Number)
    as a final safety net, and skips + reports anything that somehow
    lacks them rather than silently including a blank row.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(
            f"{path} not found. Run the scraper first (run_register.py) -- "
            f"there is nothing to build a deliverable from yet."
        )

    records_by_ref = {}
    order = []
    skipped = []

    with p.open(encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError as exc:
                skipped.append((line_no, f"malformed JSON: {exc}"))
                continue

            name = record.get("firm_details", {}).get("Name", "").strip()
            ref = record.get("firm_details", {}).get("DFSA Reference Number", "").strip()

            if not name or not ref:
                skipped.append((line_no, f"missing Name or Reference Number (name={name!r}, ref={ref!r})"))
                continue

            # De-duplicate on reference number -- if the same firm appears
            # more than once (e.g. a forced re-scrape to backfill a fixed
            # field appends a corrected record AFTER the original one, or
            # a hand-edited/manually re-run file), keep the LAST occurrence,
            # not the first. firms.jsonl is append-only, so later lines are
            # always more recent than earlier ones for the same reference
            # number -- keeping "first" would silently prefer stale/broken
            # data over a deliberate correction (confirmed real case: the
            # regulatory_actions parser was fixed after an initial full
            # run, so a re-scrape's corrected records must win).
            if ref in records_by_ref:
                skipped.append((line_no, f"duplicate DFSA Reference Number {ref}, superseded by a later line"))
            else:
                order.append(ref)
            records_by_ref[ref] = record

    records = [records_by_ref[ref] for ref in order]

    if skipped:
        print(f"NOTE: {len(skipped)} line(s) in {path} were excluded from the workbook:")
        for line_no, reason in skipped[:20]:
            print(f"  line {line_no}: {reason}")
        if len(skipped) > 20:
            print(f"  ... and {len(skipped) - 20} more")

    return records


def style_header_row(ws, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"


def autosize_columns(ws, headers: list[str], rows: list[list], max_width: int = 60) -> None:
    for col_idx, header in enumerate(headers, start=1):
        longest = len(str(header))
        for row in rows:
            val = row[col_idx - 1]
            if val is not None:
                longest = max(longest, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + 3, max_width)


def add_table(ws, headers: list[str], num_rows: int, table_name: str) -> None:
    if num_rows == 0:
        return
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{num_rows + 1}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(table)


def write_sheet(wb: Workbook, sheet_name: str, headers: list[str], rows: list[list], table_name: str):
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
    style_header_row(ws, len(headers))
    autosize_columns(ws, headers, rows)
    add_table(ws, headers, len(rows), table_name)
    return ws


def build_workbook(records: list[dict]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    # --- Sheet 1: Companies -----------------------------------------------
    # Confirmed field names present across the full 2052-firm dataset (via
    # a live key-frequency scan of firms.jsonl -- see build_deliverable.py
    # history): Name, Legal Status, DFSA Reference Number, Address,
    # Telephone Number always present; Date of Licence, Fax Number, Date of
    # Withdrawal, Endorsements, Date of Registration, DNFBP Business, Date
    # of Recognition, Exchange Membership, Audit Services, Client Assets,
    # Waivers, and four Endorsement variants present on a subset of firms.
    # Rather than hardcode that list (and silently drop any field DFSA adds
    # or renames later), the column set is the union of every firm_details
    # key actually seen in this run, so nothing collected is ever dropped
    # from the sheet -- a firm missing a given field just gets a blank
    # cell for it, same as before.
    priority_fields = ["Name", "DFSA Reference Number", "Legal Status"]
    all_field_keys: list[str] = []
    seen_keys = set()
    for r in records:
        for k in r.get("firm_details", {}).keys():
            if k not in seen_keys:
                seen_keys.add(k)
                all_field_keys.append(k)
    other_fields = sorted(k for k in all_field_keys if k not in priority_fields)
    firm_detail_fields = priority_fields + other_fields

    company_headers = firm_detail_fields + [
        "Individuals Count", "Individuals (Names & Roles)",
        "Financial Services Count", "Financial Services (Categories)",
        "Regulatory Actions Count", "Regulatory Actions (Details)",
    ]
    company_rows = []
    for r in records:
        fd = r.get("firm_details", {})
        individuals = r.get("individuals", [])
        financial_services = r.get("financial_services", [])
        regulatory_actions = r.get("regulatory_actions", [])

        individuals_text = "; ".join(
            f"{ind.get('name', '')} ({ind.get('type_of_individual', '')})".strip()
            for ind in individuals
        )
        fs_text = "; ".join(fs.get("category", "") for fs in financial_services)
        ra_text = "; ".join(
            f"{a.get('category', '')} on {a.get('date_of_use', '')}".strip()
            for a in regulatory_actions
        )

        company_rows.append(
            [fd.get(k, "") for k in firm_detail_fields]
            + [
                len(individuals),
                individuals_text,
                len(financial_services),
                fs_text,
                len(regulatory_actions),
                ra_text,
            ]
        )
    write_sheet(wb, "Companies", company_headers, company_rows, "CompaniesTable")

    # --- Sheet 2: Individuals ----------------------------------------------
    # Confirmed 100% field coverage across 32,077 individual rows (all six
    # keys present on every row): detail_url, name, reference_number,
    # type_of_individual, effective_date, date_withdrawn.
    individual_headers = [
        "Company Name", "DFSA Reference Number", "Individual Name",
        "Individual Reference Number", "Type of Individual",
        "Effective Date", "Date Withdrawn", "Individual Profile URL",
    ]
    individual_rows = []
    for r in records:
        fd = r.get("firm_details", {})
        company_name = fd.get("Name", "")
        company_ref = fd.get("DFSA Reference Number", "")
        for ind in r.get("individuals", []):
            individual_rows.append([
                company_name,
                company_ref,
                ind.get("name", ""),
                ind.get("reference_number", ""),
                ind.get("type_of_individual", ""),
                ind.get("effective_date", ""),
                ind.get("date_withdrawn", "") or "",
                ind.get("detail_url", ""),
            ])
    write_sheet(wb, "Individuals", individual_headers, individual_rows, "IndividualsTable")

    # --- Sheet 3: Financial Services (only if any exist) -------------------
    # Previously collected in every firm record but never included in the
    # workbook at all -- one row per (firm, category), with that category's
    # permitted instruments joined into a single readable cell.
    fs_rows = []
    for r in records:
        fd = r.get("firm_details", {})
        company_name = fd.get("Name", "")
        company_ref = fd.get("DFSA Reference Number", "")
        for fs in r.get("financial_services", []):
            fs_rows.append([
                company_name,
                company_ref,
                fs.get("category", ""),
                ", ".join(fs.get("instruments", []) or []),
            ])

    if fs_rows:
        fs_headers = ["Company Name", "DFSA Reference Number", "Financial Service Category", "Permitted Instruments"]
        write_sheet(wb, "Financial Services", fs_headers, fs_rows, "FinancialServicesTable")

    # --- Sheet 4: Regulatory Actions (only if any exist) -------------------
    reg_action_rows = []
    for r in records:
        fd = r.get("firm_details", {})
        company_name = fd.get("Name", "")
        company_ref = fd.get("DFSA Reference Number", "")
        for action in r.get("regulatory_actions", []):
            reg_action_rows.append([
                company_name,
                company_ref,
                action.get("title", ""),
                action.get("category", ""),
                action.get("date_of_use", ""),
                action.get("document_url", ""),
            ])

    if reg_action_rows:
        reg_action_headers = [
            "Company Name", "DFSA Reference Number", "Action / Reference",
            "Action Type", "Date", "Document URL",
        ]
        write_sheet(wb, "Regulatory Actions", reg_action_headers, reg_action_rows, "RegulatoryActionsTable")

    return wb


def organize_internal_files() -> None:
    INTERNAL_DIR.mkdir(exist_ok=True)
    for fname in (SOURCE_FIRMS_JSONL, SOURCE_ERRORS_JSONL, SOURCE_CHECKPOINT):
        src = Path(fname)
        if src.exists():
            dest = INTERNAL_DIR / fname
            shutil.copy2(src, dest)  # copy, not move -- run_register.py may still be actively writing to the originals
            print(f"Copied {fname} -> {dest}")
        else:
            print(f"NOTE: {fname} not found, skipping.")


def main():
    FINAL_DIR.mkdir(exist_ok=True)

    print(f"Reading {SOURCE_FIRMS_JSONL} ...")
    records = load_valid_records(SOURCE_FIRMS_JSONL)
    print(f"{len(records)} validated compan(y/ies) will be included in the workbook.")

    total_individuals = sum(len(r.get("individuals", [])) for r in records)
    total_reg_actions = sum(len(r.get("regulatory_actions", [])) for r in records)
    print(f"{total_individuals} individual record(s), {total_reg_actions} regulatory action record(s).")

    wb = build_workbook(records)
    wb.save(OUTPUT_XLSX)
    print(f"\nSaved: {OUTPUT_XLSX.resolve()}")

    organize_internal_files()
    print(f"\nInternal artifacts available in: {INTERNAL_DIR.resolve()}")
    print(f"Built at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()