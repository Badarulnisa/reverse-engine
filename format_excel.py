
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

input_path = "output_schemas/extracted_data.xlsx"
output_path = "output_schemas/extracted_data_formatted.xlsx"

wb = load_workbook(input_path)
ws = wb.active

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    
    for cell in col:
        if cell.row == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            
        val = str(cell.value or "")
        if len(val) > max_length:
            max_length = len(val)
            
    adjusted_width = max(max_length + 5, 18)
    if adjusted_width > 45:
        adjusted_width = 45
    ws.column_dimensions[col_letter].width = adjusted_width

ws.row_dimensions[1].height = 28
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 22

wb.save(output_path)
print(f"[+] Success! Formatted file saved to: {output_path}")

