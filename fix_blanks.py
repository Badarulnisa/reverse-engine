
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from app.core.har_parser import parse_har_file
from app.collectors.extractor import find_data_rows

har_path = "dmcc.ae.har"
records = parse_har_file(har_path)

all_rows = []
for rec in records:
    if "dmccsf.my.salesforce-sites.com" in rec.get("url", ""):
        body = rec.get("response_body")
        if body:
            rows = find_data_rows(body)
            for row in rows:
                if row:
                    clean_row = {
                        "Company Name (English)": row.get("customerName"),
                        "Company Name (Arabic)": row.get("customerArabicName"),
                        "License Number": row.get("licNo"),
                        "License Issue Date": row.get("licIssueDate"),
                        "License Expiry Date": row.get("licExpDate"),
                        "License Address (English)": str(row.get("licAddress", "")).replace("<br>", "\n"),
                        "License Address (Arabic)": str(row.get("licArabicAddress", "")).replace("<br>", "\n"),
                        "License Manager": row.get("licManagerName"),
                        "License Manager (Arabic)": row.get("licManagerArabicName"),
                        "License Activity": str(row.get("licenseActivities", "")).replace("<br>", "\n"),
                        "License Activity (Arabic)": str(row.get("licenseArabicActivities", "")).replace("<br>", "\n"),
                        "License Status": row.get("licenseStatus"),
                        "Registration Status": row.get("customerRegistrationStatus"),
                        "Registration Number": row.get("customerRegistrationNumber"),
                        "Registration Date": row.get("customerRegistrationDate")
                    }
                    all_rows.append(clean_row)

if all_rows:
    df = pd.DataFrame(all_rows)
    os.makedirs("output_schemas", exist_ok=True)
    output_path = "output_schemas/complete_company_details_v2.xlsx"
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
            
            val = str(cell.value or "")
            for line in val.split("\n"):
                if len(line) > max_length:
                    max_length = len(line)
                
        adjusted_width = max(max_length + 4, 22)
        if adjusted_width > 55:
            adjusted_width = 55
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 75

    wb.save(output_path)
    print(f"[+] Success! Fully populated records saved to: {output_path}")
else:
    print("[!] No records found.")

